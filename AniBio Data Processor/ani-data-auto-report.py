import openpyxl
import os
from dotenv import load_dotenv

translation_dict = {
    'EU Submission*': 'EU-statistik *',
    'Animal types*': 'Djurslag*',
    'Num. of animals*': 'Antal djur *',
    'Re-use*': 'Återanvändning *',
    'Genetic status*': 'Genetisk status *',
    'Creation of a new genetically altered line*': 'Framställande av en ny genetiskt förändrad stam*',
    'Purposes*': 'Syfte *',
    'Severity*': 'Svårhetsgrad *',
    'Custom severity': 'Anpassad svårhetsgrad',
    'Comment 1/Explanation of warnings': 'Kommentar 1/Förklaring av varningar',
    'Comments 2': 'Kommentar 2'
}

code_book = {
     '[Y]': '[Y] Ja',
     '[N]': '[N] Nej',
    '[A1]': '[A1] Husmus (Mus musculus)',
    '[A2]': '[A2] Brunråtta (Rattus norvegicus)',
    '[A3]': '[A3] Marsvin (Cavia porcellus)',
    '[A4]': '[A4] Guldhamster (Mesocricetus auratus)',
    '[A5]': '[A5] Kinesisk dvärghamster (Cricetulus griseus)',
    '[A6]': '[A6] Vanlig ökenråtta (gerbil) (Meriones unguiculatus)',
    '[A7]': '[A7] Andra gnagare (andra Rodentia)',
    '[A8]': '[A8] Kanin (Oryctolagus cuniculus)',
    '[A9]': '[A9] Katt (Felis catus)',
    '[A10]': '[A10] Hund (Canis familiaris)',
    '[A11]': '[A11] Frett (Mustela putorius furo)',
    '[A12]': '[A12] Andra rovdjur (andra Carnivora)',
    '[A13]': '[A13] Hästar, åsnor och korsningar (Equidae)',
    '[A14]': '[A14] Svin (Sus scrofa domesticus)',
    '[A15]': '[A15] Get (Capra aegagrus hircus)',
    '[A16]': '[A16] Får (Ovis aries)',
    '[A17]': '[A17] Nötkreatur (Bos taurus)',
    '[A18]': '[A18] Halvapor (Prosimia)',
    '[A19]': '[A19] Silkesapor och tamariner (t.ex. Callithrix jacchus)',
    '[A20]': '[A20] Krabbmakak (Macaca fascicularis)',
    '[A21]': '[A21] Rhesusmakak (Macaca mulatta)',
    '[A22]': '[A22] Gröna markattor (Chlorocebus spp.) (vanligtvis antingen pygerythrus eller sabaeus)',
    '[A23]': '[A23] Babianer (Papio spp.)',
    '[A24]': '[A24] Dödskalleapor (t.ex. Saimiri sciureus)',
    '[A25-1]': '[A25-1] Andra arter av östapor (andra arter av Cercopithecoidea)',
    '[A25-2]': '[A25-2] Andra arter av västapor (andra arter av Ceboidea)',
    '[A26]': '[A26] Människoartade primater (Hominoidea)',
    '[A27]': '[A27] Andra däggdjur (andra Mammalia)',
    '[A28]': '[A28] Tamhöna (Gallus gallus domesticus)',
    '[A37]': '[A37] Kalkon (Meleagris gallopavo)',
    '[A29]': '[A29] Andra fåglar (andra Aves)',
    '[A30]': '[A30] Kräldjur (Reptilia)',
    '[A31]': '[A31] Grodor (Rana temporaria och Rana pipiens)',
    '[A32]': '[A32] Klogrodor (Xenopus laevis och Xenopus tropicalis)',
    '[A33]': '[A33] Andra groddjur (andra Amphibia)',
    '[A34]': '[A34] Sebrafisk (Danio rerio)',
    '[A35]': '[A35] Andra fiskar (andra Pisces)',
    '[A36]': '[A36] Bläckfiskar (Cephalopoda)',
    '[A38]': '[A38] Havsabborrfiskar (arter från familjer som t.ex. Serranidae, Moronidae)',
    '[A39]': '[A39] Lax, öring, röding och harr (Salmonidae)',
    '[A40]': '[A40] Levandefödande tandkarpar (Poeciliidae)',
    '[GS1]': '[GS1] Icke genetiskt förändrad',
    '[GS2]': '[GS2] Genetiskt förändrad utan skadlig fenotyp',
    '[GS3]': '[GS3] Genetiskt förändrad med skadlig fenotyp',
    '[SV1]': '[SV1] Terminal',
    '[SV2]': '[SV2] Ringa (upp till och med)',
    '[SV3]': '[SV3] Måttlig',
    '[SV4]': '[SV4] Avsevärd',
    '[IG1]': '[IG1] Invasiv genotypning: blodprov',
    '[IG2]': '[IG2] Invasiv genotypning: öronbiopsi',
    '[IG3]': '[IG3] Invasiv genotypning: svansbiopsi',
    '[IG6]': '[IG6] Invasiv genotypning: fenbiopsi',
    '[IG4]': '[IG4] Invasiv genotypning: tåklippning',
    '[IG5]': '[IG5] Invasiv genotypning: andra metoder',
    '[ST1]': '[ST1] Överbliven vävnad från märkning av djur genom öronklippning',
    '[ST2]': '[ST2] Överbliven vävnad från märkning av djur genom tåklippning',
    '[NG1]': '[NG1] Ickeinvasiv genotypning: hårprover',
    '[NG2]': '[NG2] Ickeinvasiv genotypning: observation under särskilt ljus',
    '[NG3]': '[NG3] Ickeinvasiv genotypning: post mortem',
    '[NG4]': '[NG4] Ickeinvasiv genotypning: andra metoder'
}

def process_column(header):
    if header in translation_dict:
        return translation_dict[header]
    else:
        return header

def process_code_book(cell_value):
    if isinstance(cell_value, int):
        return cell_value
    elif isinstance(cell_value, str):
        for code, translation in code_book.items():
            if code in cell_value:
                return translation
        return cell_value
    else:
        return cell_value

def process_excel(source_file: str, target_file: str):
    source_wb = openpyxl.load_workbook(source_file)
    target_wb = openpyxl.load_workbook(target_file)
    
    source_ws = source_wb.active
    target_ws = target_wb.active
    
    column_mapping = {}
    for col in range(1, source_ws.max_column + 1):
        header = source_ws.cell(row=1, column=col).value
        translated_header = process_column(header)
        if translated_header in translation_dict.values():
            column_mapping[col] = translated_header
    
    for col in column_mapping:
        source_col = col
        target_col = None
        for col_idx in range(1, target_ws.max_column + 1):
            if target_ws.cell(row=1, column=col_idx).value == column_mapping[source_col]:
                target_col = col_idx
                break
        if target_col is not None:
            for row in range(2, source_ws.max_row + 1):
                cell_value = source_ws.cell(row=row, column=source_col).value
                translated_value = process_code_book(cell_value)
                target_ws.cell(row=row, column=target_col, value=translated_value)
    
    target_wb.save(target_file)
    print("Data proccess successful")

if __name__ == "__main__":
    load_dotenv()
    process_excel(os.getenv("SOURCE_FILE"), os.getenv("TEMPLATE_FILE"))